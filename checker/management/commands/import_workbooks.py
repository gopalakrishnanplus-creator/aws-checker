from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils.text import slugify
from openpyxl import load_workbook

from checker.importers import sync_check_definitions, sync_resources

CHECK_CODE_OVERRIDES = {
    "4xx errors": "client_4xx_errors",
    "5xx errors": "server_5xx_errors",
}


class Command(BaseCommand):
    help = "Import AWS check definitions and resource inventories from workbook exports."

    def add_arguments(self, parser):
        parser.add_argument("--details", required=True, help="Path to AWS availability check details workbook")
        parser.add_argument("--ec2", required=True, help="Path to EC2 inventory workbook")
        parser.add_argument("--rds", required=True, help="Path to RDS inventory workbook")
        parser.add_argument("--s3", required=True, help="Path to S3 inventory workbook")
        parser.add_argument("--account-id", default="736616688306", help="AWS account ID to stamp on imported resources")

    def handle(self, *args, **options):
        paths = {key: Path(options[key]) for key in ("details", "ec2", "rds", "s3")}
        for label, path in paths.items():
            if not path.exists():
                raise CommandError(f"{label} workbook not found: {path}")

        definitions = self._parse_check_details(paths["details"])
        resources = []
        resources.extend(self._parse_ec2(paths["ec2"], options["account_id"]))
        resources.extend(self._parse_rds(paths["rds"], options["account_id"]))
        resources.extend(self._parse_s3(paths["s3"], options["account_id"]))

        definition_count = sync_check_definitions(definitions, replace=True)
        resource_count = sync_resources(resources, replace=True)
        self.stdout.write(
            self.style.SUCCESS(
                f"Imported {definition_count} check definitions and {resource_count} resources from workbooks."
            )
        )

    def _parse_check_details(self, path):
        workbook = load_workbook(path, data_only=True)
        mapping = {
            "EC2_Checks": "ec2",
            "RDS_Checks": "rds",
            "S3_Checks": "s3",
        }
        results = {}
        for sheet_name, service_type in mapping.items():
            worksheet = workbook[sheet_name]
            definitions = []
            for sort_order, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=1):
                if not any(row):
                    continue
                definitions.append(
                    {
                        "sort_order": sort_order,
                        "category": row[0],
                        "code": CHECK_CODE_OVERRIDES.get(row[1], slugify(row[1]).replace("-", "_")),
                        "check_item": row[1],
                        "what_to_verify": row[2],
                        "how_to_check": row[3],
                        "success_criteria": row[4],
                        "priority": row[5],
                        "frequency": row[6],
                    }
                )
            results[service_type] = definitions
        return results

    def _parse_ec2(self, path, account_id):
        worksheet = load_workbook(path, data_only=True).active
        resources = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            region = (row[3] or "")[:-1]
            security_groups = [item.strip() for item in (row[7] or "").split(",") if item and item.strip()]
            resources.append(
                {
                    "service_type": "ec2",
                    "account_id": account_id,
                    "name": row[0],
                    "resource_identifier": row[1],
                    "region": region,
                    "availability_zone": row[3] or "",
                    "endpoint": row[4] or "",
                    "public_ip_address": row[5] or None,
                    "elastic_ip": row[6] or None,
                    "metadata": {
                        "instance_type": row[2],
                        "security_groups": security_groups,
                        "platform_details": row[8],
                    },
                    "check_config": {},
                }
            )
        return resources

    def _parse_rds(self, path, account_id):
        worksheet = load_workbook(path, data_only=True).active
        resources = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            engine = str(row[2])
            port = 5432 if "postgres" in engine.lower() else 3306
            resources.append(
                {
                    "service_type": "rds",
                    "account_id": account_id,
                    "name": row[0],
                    "resource_identifier": row[0],
                    "region": row[9],
                    "availability_zone": row[10] or "",
                    "endpoint": row[1],
                    "port": port,
                    "engine": engine,
                    "resource_state": row[4] or "",
                    "metadata": {
                        "version": str(row[3]),
                        "instance_class": row[5],
                        "vcpu": int(row[6]) if row[6] is not None else None,
                        "ram": row[7],
                        "storage": row[8],
                        "multi_az": str(row[11]).lower() == "yes",
                        "vpc": row[12],
                        "seed_cpu": row[13],
                        "seed_connections": row[14],
                    },
                    "check_config": {"db_probe": {"username_env": "", "password_env": "", "database_env": "", "port": port}},
                }
            )
        return resources

    def _parse_s3(self, path, account_id):
        worksheet = load_workbook(path, data_only=True).active
        resources = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            resources.append(
                {
                    "service_type": "s3",
                    "account_id": account_id,
                    "name": row[0],
                    "resource_identifier": row[0],
                    "region": row[1],
                    "endpoint": row[2],
                    "metadata": {
                        "created_on": row[3].date().isoformat() if row[3] else "",
                        "versioning": str(row[4]).replace("❌ ", ""),
                        "mfa_delete": str(row[5]).replace("❌ ", ""),
                        "abac": str(row[6]).replace("❌ ", ""),
                    },
                    "check_config": {"canary_prefix": f"aws-checker/probes/{row[0]}"},
                }
            )
        return resources
